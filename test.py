from openpyxl import load_workbook




























book = load_workbook(r'\\172.16.8.87\d\Dauren\КЕКУС.xlsx')
book1 = load_workbook(r'\\172.16.8.87\d\Dinara\Аренда выгрузка с БИНами.xlsx')

sheet = book.active
sheet1 = book1.active

for row in range(6, sheet.max_row + 1):
    for row1 in range(2, sheet1.max_row + 1):

        bin_ = sheet[f'A{row}'].value == sheet1[f'D{row1}'].value
        name1 = sheet[f'C{row}'].value == sheet1[f'A{row1}'].value
        name2 = sheet[f'E{row}'].value == sheet1[f'D{row1}'].value

        if bin_ or name1 or name2:
            sheet[f'M{row}'].value = sheet1[f'B{row1}'].value

book.save('prrr.xlsx')
